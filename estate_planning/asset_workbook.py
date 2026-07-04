"""Multi-tab Excel asset sheet: a template to fill offline (one tab per asset
category, with category-specific columns and mandatory fields marked) and a
parser for the uploaded workbook.
"""

import io

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .asset_schema import ASSET_FIELD_SCHEMA
from .models import ASSET_CATEGORIES, Asset

DATA_START_ROW = 3  # row 1 = English labels, row 2 = Thai labels, row 3+ = data

REQUIRED_FILL = PatternFill("solid", fgColor="F8CBAD")  # peach = required
OPTIONAL_FILL = PatternFill("solid", fgColor="E7E6E6")  # grey = optional
HEADER_FONT = Font(bold=True)
TH_FONT = Font(italic=True, size=9, color="666666")

# Excel sheet titles must avoid \ / ? * [ ] : and be <= 31 chars. Keep an explicit
# safe title per category, used for both generation and parsing.
SHEET_TITLES = {
    "real_estate": "Real estate",
    "vehicle": "Vehicle",
    "bank_account": "Bank account",
    "investment": "Investment",
    "insurance": "Insurance",
    "business": "Business",
    "digital": "Digital asset",
    "other": "Other",
}
_TITLE_TO_CATEGORY = {title: cat for cat, title in SHEET_TITLES.items()}


def _instructions(ws):
    ws.column_dimensions["A"].width = 100
    lines = [
        ("How to use this asset sheet / วิธีใช้บัญชีทรัพย์สิน", True),
        ("", False),
        ("• Each tab below is one asset category. Fill in only the tabs that apply to you.", False),
        ("  แต่ละแท็บคือทรัพย์สินหนึ่งประเภท กรอกเฉพาะแท็บที่เกี่ยวข้องกับท่าน", False),
        ("• Columns marked * and shaded peach are REQUIRED if you list that asset.", False),
        ("  Grey columns are optional. / คอลัมน์ที่มี * และแรเงาสีส้มคือช่องบังคับ ช่องสีเทาเป็นทางเลือก", False),
        ("• Enter values in Thai Baht (THB). For land and buildings use the official", False),
        ("  appraised value ราคาประเมิน from https://assessprice.treasury.go.th", False),
        ("• For digital assets, do NOT enter real passwords or seed phrases — only note", False),
        ("  where those access details can be found. / อย่ากรอกรหัสผ่านจริง", False),
        ("• When finished, upload this file back into the app.", False),
        ("", False),
        ("Legend: peach = required, grey = optional.", False),
    ]
    for i, (text, is_title) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        if is_title:
            c.font = Font(bold=True, size=14)


def template_xlsx():
    """Return the .xlsx template as bytes."""
    wb = openpyxl.Workbook()
    _instructions(wb.active)
    wb.active.title = "Instructions"

    for cat in ASSET_CATEGORIES:
        ws = wb.create_sheet(title=SHEET_TITLES[cat])
        for col, (key, en, th, required) in enumerate(
            ASSET_FIELD_SCHEMA[cat]["fields"], start=1
        ):
            top = ws.cell(row=1, column=col, value=en + (" *" if required else ""))
            top.font = HEADER_FONT
            top.fill = REQUIRED_FILL if required else OPTIONAL_FILL
            top.alignment = Alignment(wrap_text=True, vertical="center")
            th_cell = ws.cell(row=2, column=col, value=th)
            th_cell.font = TH_FONT
            th_cell.alignment = Alignment(wrap_text=True, vertical="center")
            ws.column_dimensions[get_column_letter(col)].width = max(18, len(en) + 2)
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _to_float(raw):
    if raw is None:
        return 0.0
    if isinstance(raw, (int, float)):
        return float(raw)
    try:
        return float(str(raw).replace(",", "").strip())
    except ValueError:
        return 0.0


def parse_workbook(data_bytes):
    """Parse an uploaded .xlsx asset sheet into a list of Asset."""
    wb = openpyxl.load_workbook(io.BytesIO(data_bytes), data_only=True)
    assets = []
    for ws in wb.worksheets:
        cat = _TITLE_TO_CATEGORY.get(ws.title.strip())
        if not cat:
            continue  # e.g. the Instructions tab
        schema = ASSET_FIELD_SCHEMA[cat]
        keys = [f[0] for f in schema["fields"]]
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            values = list(row)[: len(keys)]
            if not any(v not in (None, "") for v in values):
                continue
            rowdict = {
                k: v for k, v in zip(keys, values) if v not in (None, "")
            }
            details = {
                k: str(v).strip()
                for k, v in rowdict.items()
                if k not in ("value_thb", "notes") and str(v).strip()
            }
            assets.append(
                Asset(
                    category=cat,
                    description=str(rowdict.get(schema["primary"], "") or "").strip(),
                    value_thb=_to_float(rowdict.get("value_thb")),
                    location=str(rowdict.get(schema["reference"], "") or "").strip(),
                    notes=str(rowdict.get("notes", "") or "").strip(),
                    details=details,
                )
            )
    return assets
