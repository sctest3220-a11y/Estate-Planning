import io
import unittest

import openpyxl

from estate_planning.asset_schema import ASSET_FIELD_SCHEMA
from estate_planning.asset_workbook import (
    DATA_START_ROW,
    SHEET_TITLES,
    parse_workbook,
    template_xlsx,
)
from estate_planning.models import ASSET_CATEGORIES


class TemplateTests(unittest.TestCase):
    def setUp(self):
        self.data = template_xlsx()
        self.wb = openpyxl.load_workbook(io.BytesIO(self.data))

    def test_has_instructions_and_a_tab_per_category(self):
        self.assertIn("Instructions", self.wb.sheetnames)
        for cat in ASSET_CATEGORIES:
            self.assertIn(SHEET_TITLES[cat], self.wb.sheetnames)

    def test_required_fields_marked_with_asterisk(self):
        ws = self.wb["Bank account"]
        headers = [c.value for c in ws[1]]
        # bank_name and account_no are required
        self.assertTrue(any(h and h.startswith("Bank name") and h.endswith("*") for h in headers))
        self.assertTrue(any(h and h.startswith("Account number") and h.endswith("*") for h in headers))
        # account_type is optional -> no asterisk
        self.assertTrue(any(h == "Account type (savings / current / fixed)" for h in headers))

    def test_thai_labels_in_second_row(self):
        ws = self.wb["Bank account"]
        row2 = [c.value for c in ws[2]]
        self.assertIn("ชื่อธนาคาร", row2)

    def test_sheet_titles_are_valid(self):
        for title in SHEET_TITLES.values():
            self.assertLessEqual(len(title), 31)
            self.assertFalse(any(ch in title for ch in "\\/?*[]:"))


class ParseTests(unittest.TestCase):
    def _filled_workbook(self):
        wb = openpyxl.load_workbook(io.BytesIO(template_xlsx()))
        bank = wb["Bank account"]
        bank.cell(row=DATA_START_ROW, column=1, value="Bangkok Bank")
        bank.cell(row=DATA_START_ROW, column=2, value="123-4-56789")
        bank.cell(row=DATA_START_ROW, column=3, value="Savings")
        bank.cell(row=DATA_START_ROW, column=6, value=850000)
        re_ = wb["Real estate"]
        re_.cell(row=DATA_START_ROW, column=1, value="Condo")
        re_.cell(row=DATA_START_ROW, column=6, value=6500000)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_parses_rows_across_tabs(self):
        assets = parse_workbook(self._filled_workbook())
        cats = sorted(a.category for a in assets)
        self.assertEqual(cats, ["bank_account", "real_estate"])

    def test_details_captured(self):
        assets = parse_workbook(self._filled_workbook())
        bank = next(a for a in assets if a.category == "bank_account")
        self.assertEqual(bank.description, "Bangkok Bank")
        self.assertEqual(bank.location, "123-4-56789")
        self.assertEqual(bank.value_thb, 850000)
        self.assertEqual(bank.details.get("account_type"), "Savings")

    def test_blank_rows_skipped(self):
        # An untouched template has no data rows.
        assets = parse_workbook(template_xlsx())
        self.assertEqual(assets, [])

    def test_instructions_tab_ignored(self):
        # No sheet titled by a real category maps to Instructions.
        assets = parse_workbook(template_xlsx())
        self.assertEqual(assets, [])

    def test_schema_covers_all_categories(self):
        self.assertEqual(set(ASSET_FIELD_SCHEMA), set(ASSET_CATEGORIES))


if __name__ == "__main__":
    unittest.main()
